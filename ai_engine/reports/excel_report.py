import sys
import os
from pathlib import Path
from typing import Dict, Any, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ensure ai_engine root is in sys.path
sys.path.append(str(Path(__file__).resolve().parent.parent))

from utils.logger import get_logger

logger = get_logger(__name__)

class ExcelReportBuilder:
    """
    Generates structured Excel workbooks using OpenPyXL.
    Consumes centralized AI results without duplicating prediction logic.
    """

    @staticmethod
    def generate_student_excel_report(
        output_filepath: str,
        feature_dict: Dict[str, Any],
        prediction_dict: Dict[str, Any],
        pattern_dict: Dict[str, Any],
        recommendations: List[Dict[str, Any]],
        xai_dict: Dict[str, Any]
    ) -> str:
        """
        Builds a styled multi-worksheet Excel workbook for student attendance analytics.
        """
        os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
        wb = Workbook()

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

        # Worksheet 1: Executive Summary & Predictions
        ws_summary = wb.active
        ws_summary.title = "Summary & Predictions"

        ws_summary.append(["Student Attendance AI Report"])
        ws_summary.append(["Student ID", feature_dict.get("student_id", 0)])
        ws_summary.append(["Executive Summary", xai_dict.get("explanation_text", "")])
        ws_summary.append([])

        headers_pred = ["Metric", "Value"]
        ws_summary.append(headers_pred)
        for col_num, h in enumerate(headers_pred, 1):
            cell = ws_summary.cell(row=5, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        from config import settings
        from models.forecasting import AttendanceForecaster
        req_pct = settings.REQUIRED_ATTENDANCE_PCT
        classes_att = feature_dict.get("classes_attended", 0)
        total_cond = feature_dict.get("total_conducted", 0)
        forecast_info = AttendanceForecaster.calculate_classes_needed(classes_att, total_cond, target_pct=req_pct)
        classes_needed_str = str(forecast_info.get("classes_needed", 0)) if not forecast_info.get("is_target_achieved") else "0 (Requirement Met)"
        if not forecast_info.get("is_target_achieved") and not forecast_info.get("is_achievable"):
            classes_needed_str = "Unreachable with remaining classes"

        cal_source = pattern_dict.get("source_type", feature_dict.get("calendar_source", "SYNTHETIC_DEVELOPMENT"))
        cal_source_label = "Uploaded PDF" if cal_source == "UPLOADED_PDF" else "SYNTHETIC_DEVELOPMENT"

        pred_rows = [
            ("Current Attendance Percentage", f"{feature_dict.get('current_attendance_pct', 0.0)}%"),
            ("Required Attendance Criterion", f"{req_pct}%"),
            ("Predicted Semester Percentage", f"{prediction_dict.get('predicted_pct', 0.0)}%"),
            ("Prediction Interval Bounds", f"[{prediction_dict.get('predicted_min_pct', 0.0)}% - {prediction_dict.get('predicted_max_pct', 100.0)}%]"),
            ("Status vs Requirement", "Below Requirement" if prediction_dict.get("predicted_pct", 0.0) < req_pct else "Meets Requirement"),
            ("Classes Needed for Requirement", classes_needed_str),
            ("Risk Category", prediction_dict.get("risk_level", "LOW")),
            ("Prediction Reliability", prediction_dict.get("prediction_reliability", "MEDIUM")),
            ("Academic Calendar Source", cal_source_label),
            ("Total Conducted Classes", total_cond),
            ("Attended Classes", classes_att),
            ("Absent Classes", feature_dict.get("classes_absent", 0)),
            ("Remaining Classes", feature_dict.get("remaining_classes", 0))
        ]
        for row in pred_rows:
            ws_summary.append(list(row))

        # Worksheet 2: Patterns & Recommendations
        ws_patterns = wb.create_sheet(title="Patterns & Recommendations")
        ws_patterns.append(["Pattern Analysis Feature", "Metric Value"])
        for col_num, h in enumerate(["Pattern Analysis Feature", "Metric Value"], 1):
            cell = ws_patterns.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        worst_sub_name = feature_dict.get("worst_subject_name", f"Subject ID {pattern_dict.get('worst_performing_subject_id', 0)}")

        pattern_rows = [
            ("Most Absent Weekday", pattern_dict.get("most_absent_weekday", "N/A")),
            ("Worst Weekday Absence Rate", f"{pattern_dict.get('worst_day_absence_rate', 0.0)}%"),
            ("Worst Subject", worst_sub_name),
            ("Worst Subject Attendance %", f"{pattern_dict.get('worst_subject_attendance_pct', 0.0)}%"),
            ("Pre-Holiday Drop %", f"{pattern_dict.get('pre_holiday_drop_pct', 0.0)}%"),
            ("Trend Direction", pattern_dict.get("trend_direction", "STABLE"))
        ]
        for row in pattern_rows:
            ws_patterns.append(list(row))

        ws_patterns.append([])
        ws_patterns.append(["Priority", "Recommendation Text"])
        row_start = len(pattern_rows) + 3
        for col_num, h in enumerate(["Priority", "Recommendation Text"], 1):
            cell = ws_patterns.cell(row=row_start, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        for rec in recommendations:
            ws_patterns.append([rec.get("priority", "MEDIUM"), rec.get("recommendation_text", "")])

        # Auto-adjust column widths
        for ws in [ws_summary, ws_patterns]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        wb.save(output_filepath)
        logger.info(f"Successfully generated student Excel report at {output_filepath}")
        return output_filepath
